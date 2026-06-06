# =========================================================
# ABC GAMING - SCRAPER FINAL
# Versión: merge completo + corrección de monedas
# =========================================================
#
# ✔ Todas las tiendas (v1 + v2 unificadas)
# ✔ Eneba completo
# ✔ Tiendas .com.ar → moneda forzada ARS
# ✔ Campo "moneda_tienda" por tienda
# ✔ Conversión USD / ARS / UYU
# ✔ Mejores precios por juego
# ✔ Hojas separadas por plataforma
# ✔ Links clickeables
# ✔ Formato visual dark header
# ✔ Deduplicación
#
# =========================================================
import re
import asyncio
import requests
import pandas as pd
from bs4 import BeautifulSoup
from playwright.async_api import async_playwright
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# =========================================================
# CONFIG GLOBAL
# =========================================================
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 Chrome/124 Safari/537.36"
    )
}
PAGINAS     = 2
USD_TO_UYU  = 40
ARS_TO_UYU  = 0.035

# =========================================================
# MONEDAS POR TIENDA
#
# Si una tienda no está aquí, se detecta automáticamente
# del texto del precio. Las tiendas .com.ar siempre ARS.
# =========================================================
MONEDA_TIENDA = {
    # Shopify — precios en USD
    "Alpha PS5 Primaria":       "USD",
    "Alpha PS5 Secundaria":     "USD",
    "Alpha PS4 Primaria":       "USD",
    "Alpha PS4 Secundaria":     "USD",
    "Alpha Xbox":               "USD",
    "GamerLab PS4":             "USD",
    "GamerLab PS5":             "USD",
    "GamerLab EA Play":         "USD",
    "GamerLab PSN":             "USD",
    "GamerLab All":             "USD",
    # WooCommerce — precios en UYU
    "DigitalWorld PS4":         "UYU",
    "DigitalWorld PS5":         "UYU",
    "DigitalWorld VR":          "UYU",
    "DigitalWorld Xbox":        "UYU",
    "DigitalWorld Switch":      "UYU",
    "UruguayDigital PS3":       "UYU",
    "UruguayDigital PS4":       "UYU",
    "UruguayDigital PS5":       "UYU",
    "UruguayDigital Switch1":   "UYU",
    "UruguayDigital Switch2":   "UYU",
    "UruguayDigital Xbox One":  "UYU",
    "UruguayDigital Xbox XS":   "UYU",
    "UruguayDigital PSN Plus":  "UYU",
    "UruguayDigital Nintendo":  "UYU",
    "UruguayDigital GamePass":  "UYU",
    "UYJuegos PS4":             "UYU",
    "UYJuegos PS4 VR":          "UYU",
    "UYJuegos PS5":             "UYU",
    "UYJuegos PC":              "UYU",
    "UYJuegos Xbox XS":         "UYU",
    "UYJuegos Xbox One":        "UYU",
    "UYJuegos Nintendo":        "UYU",
    "UYJuegos Nintendo2":       "UYU",
    "UYJuegos GiftCard Amazon": "UYU",
    "UYJuegos GiftCard iTunes": "UYU",
    "WebGame PS4":              "UYU",
    "WebGame PS5":              "UYU",
    "WebGame GiftCards":        "UYU",
    # Tiendas argentinas — ARS
    "PortalGames PS3":          "ARS",
    "PortalGames PS4":          "ARS",
    "PortalGames PS5":          "ARS",
    "PortalGames Switch":       "ARS",
    "MDQ PS5":                  "ARS",
    "MDQ PS4":                  "ARS",
    "MDQ PS3":                  "ARS",
    "MDQ PS Plus":              "ARS",
    "MDQ PSN Card":             "ARS",
    "MDQ Fortnite":             "ARS",
    "PlayX PS4":                "ARS",
    "PlayX PS5":                "ARS",
    "PlayX Xbox One":           "ARS",
    "PlayX Xbox XS":            "ARS",
    "PlayX Switch":             "ARS",
    "PlayX Switch 2":           "ARS",
    "PlayX Steam":              "ARS",
    "PlayX PSN Plus":           "ARS",
    "PlayX GamePass":           "ARS",
    "PlayX Switch Online":      "ARS",
    "TodoDigital PS4":          "ARS",
    "TodoDigital PS5":          "ARS",
    "TodoDigital Switch":       "ARS",
    "JDP4P5 PS4":               "ARS",
    "JDP4P5 PS5":               "ARS",
    "JDP4P5 Switch":            "ARS",
    "JDP4P5 PS Plus":           "ARS",
    "JDP4P5 GamePass":          "ARS",
    "JDP4P5 Nintendo Online":   "ARS",
    # Dix — precios en UYU
    "Dix Fortnite":             "UYU",
    "Dix PS5":                  "UYU",
    "Dix PS4":                  "UYU",
    "Dix PS3":                  "UYU",
    "Dix FC Points":            "UYU",
    "Dix Nintendo":             "UYU",
    "Dix Playstation":          "UYU",
    "Dix Razer":                "UYU",
    "Dix Steam":                "UYU",
    "Dix Xbox":                 "UYU",
    # EstacionPlay — precios en ARS (tienda argentina)
    "EstacionPlay PS4":         "ARS",
    "EstacionPlay PS5":         "ARS",
    "EstacionPlay Memb":        "ARS",
    # ZonaDigital — precios en UYU
    "ZonaDigital PS4":          "UYU",
    "ZonaDigital PS5":          "UYU",
    "ZonaDigital PC Codigo":    "UYU",
    "ZonaDigital PC Cuenta":    "UYU",
    "ZonaDigital Xbox":         "UYU",
    "ZonaDigital Apex":         "UYU",
    "ZonaDigital Fortnite VB":  "UYU",
    "ZonaDigital Fortnite":     "UYU",
    "ZonaDigital GiftCard PSN": "UYU",
    "ZonaDigital GiftCard Steam": "UYU",
    "ZonaDigital Roblox":       "UYU",
    "ZonaDigital GiftCard Xbox": "UYU",
    "ZonaDigital BattleNet":    "UYU",
    "ZonaDigital Riot":         "UYU",
    "ZonaDigital FreeFire":     "UYU",
    "ZonaDigital Google Play":  "UYU",
    "ZonaDigital Apple":        "UYU",
    "ZonaDigital Switch":       "UYU",
    "ZonaDigital Switch USA":   "UYU",
    "ZonaDigital GamePass":     "UYU",
    "ZonaDigital EA Play":      "UYU",
    # Eneba — USD
    "Eneba":                    "USD",
}

# =========================================================
# NORMALIZACION
# =========================================================
BASURA_NOMBRES = [
    "DIGITAL", "PRIMARIO", "SECUNDARIO", "CUENTA",
    "OFFLINE", "ONLINE", "PS4", "PS5", "PS3",
    "XBOX", "NINTENDO", "STANDARD", "DELUXE", "ULTIMATE",
    "LATAM", "GLOBAL", "STEAM KEY", "KEY", "PC",
    "CODIGO", "CÓDIGO", "ACCOUNT"
]

def normalizar_nombre(nombre):
    nombre = str(nombre).upper()
    for b in BASURA_NOMBRES:
        nombre = nombre.replace(b, "")
    return re.sub(r"\s+", " ", nombre).strip()

def limpiar_nombre_eneba(nombre):
    remover = [
        "Código de Steam", "Steam Key", "Steam", "GLOBAL", "LATAM", "EUROPE",
        "United States", "UNITED STATES", "Xbox Live", "Xbox", "PSN", "Nintendo",
        "Código de", "Código", "Key",
    ]
    for r in remover:
        nombre = nombre.replace(r, "")
    return re.sub(r"\s+", " ", nombre).strip()

# =========================================================
# MONEDAS Y PRECIOS
# =========================================================
def detectar_moneda_texto(precio_txt):
    """Detecta moneda del texto del precio."""
    txt = str(precio_txt).upper()
    if "USD" in txt or "US$" in txt:
        return "USD"
    if "ARS" in txt:
        return "ARS"
    if "$U" in txt or "UYU" in txt:
        return "UYU"
    return None  # No detectado — se usará moneda_tienda como fallback

def resolver_moneda(precio_txt, nombre_tienda):
    """
    Prioridad:
    1. Si el texto del precio tiene moneda explícita → usar esa.
    2. Si la URL de la tienda es .com.ar → ARS.
    3. Si la tienda tiene entrada en MONEDA_TIENDA → usar esa.
    4. Fallback: UYU.
    """
    moneda_texto = detectar_moneda_texto(precio_txt)
    if moneda_texto:
        return moneda_texto
    tienda_moneda = MONEDA_TIENDA.get(nombre_tienda)
    if tienda_moneda:
        return tienda_moneda
    return "UYU"

def limpiar_precio(precio_txt, nombre_tienda=None):
    txt    = str(precio_txt)
    moneda = resolver_moneda(txt, nombre_tienda or "")
    numero = re.search(r"[\d.,]+", txt)
    if not numero:
        return None, moneda

    valor_txt = numero.group().strip()

    if moneda == "USD":
        if "," in valor_txt and "." not in valor_txt:
            valor_txt = valor_txt.replace(",", ".")
        elif "." in valor_txt and "," in valor_txt:
            valor_txt = valor_txt.replace(",", "")
        try:
            return float(valor_txt), moneda
        except:
            return None, moneda
    else:
        valor_txt = valor_txt.replace(".", "")
        valor_txt = valor_txt.replace(",", ".")
        try:
            return float(valor_txt), moneda
        except:
            return None, moneda

def convertir_uyu(valor, moneda):
    if valor is None:
        return None
    if moneda == "USD":
        return round(valor * USD_TO_UYU, 2)
    if moneda == "ARS":
        return round(valor * ARS_TO_UYU, 2)
    return round(valor, 2)

def extraer_precio_eneba(texto):
    texto = texto.replace("Desde", "").replace("From", "").strip()
    match = re.search(r'(\d+[.,]?\d*)\s*US\$', texto)
    if match:
        try:
            return f"{float(match.group(1).replace(',', '.')):.2f} USD"
        except:
            pass
    match2 = re.search(r'\$\s*(\d+[.,]\d+)', texto)
    if match2:
        try:
            return f"{float(match2.group(1).replace(',', '.')):.2f} USD"
        except:
            pass
    return ""

# =========================================================
# DETECTAR PLATAFORMA
# =========================================================
def detectar_plataforma(nombre, tienda):
    txt = f"{nombre} {tienda}".upper()
    if "PS5" in txt:
        return "PS5"
    if "PS4" in txt:
        return "PS4"
    if "PS3" in txt:
        return "PS3"
    if "XBOX" in txt or "GAME PASS" in txt or "GAMEPASS" in txt:
        return "Xbox"
    if "NINTENDO" in txt or "SWITCH" in txt:
        return "Nintendo"
    if "STEAM" in txt:
        return "Steam"
    if "FORTNITE" in txt or "PAVOS" in txt or "V-BUCK" in txt:
        return "Fortnite"
    if "ROBLOX" in txt or "ROBUX" in txt:
        return "Robux"
    if "FREE FIRE" in txt or "FREEFIRE" in txt:
        return "Free Fire"
    if "GENSHIN" in txt:
        return "Genshin"
    if "VALORANT" in txt or "RIOT" in txt:
        return "Valorant"
    if "PUBG" in txt:
        return "PUBG"
    if "FC POINT" in txt or "FIFA" in txt or "FUT" in txt:
        return "FC Points"
    if "GTA" in txt or "SHARK" in txt:
        return "GTA"
    if "EA PLAY" in txt or "ORIGIN" in txt:
        return "EA Play"
    if "UBISOFT" in txt or "UPLAY" in txt:
        return "Ubisoft"
    if "EPIC" in txt:
        return "Epic Games"
    if "GOG" in txt:
        return "GOG"
    if "BATTLE" in txt or "BLIZZARD" in txt:
        return "BattleNet"
    if "DISCORD" in txt:
        return "Discord"
    if "TWITCH" in txt:
        return "Twitch"
    if "RAZER" in txt:
        return "Razer Gold"
    if "GOOGLE PLAY" in txt:
        return "Google Play"
    if "PSN" in txt or "PS PLUS" in txt or "PLAYSTATION PLUS" in txt:
        return "PSN Membresia"
    if "GIFT" in txt or "CARD" in txt or "GIFTCARD" in txt:
        return "Gift Cards"
    if "APEX" in txt:
        return "Apex Legends"
    return "Otros"

# =========================================================
# SCRAPERS
# =========================================================
def scrape_shopify(url, nombre_tienda, paginas=PAGINAS):
    productos = []
    moneda_forzada = MONEDA_TIENDA.get(nombre_tienda, "USD")
    for page in range(1, paginas + 1):
        page_url = f"{url}/products.json?limit=250&page={page}"
        try:
            r    = requests.get(page_url, headers=HEADERS, timeout=30)
            data = r.json()
            items = data.get("products", [])
            if not items:
                break
            for p in items:
                try:
                    precio_raw = float(p["variants"][0]["price"])
                    precio_uyu = convertir_uyu(precio_raw, moneda_forzada)
                    productos.append({
                        "nombre":          p["title"],
                        "precio_original": f"{precio_raw} {moneda_forzada}",
                        "precio":          precio_raw,
                        "moneda":          moneda_forzada,
                        "precio_uyu":      precio_uyu,
                        "link":            f"{url.split('/collections')[0]}/products/{p['handle']}"
                    })
                except:
                    continue
        except:
            break
    return productos

def scrape_woocommerce(url, nombre_tienda, paginas=PAGINAS):
    productos = []
    selectors = [
        "li.product", "div.product", ".product-small",
        ".type-product", ".product-grid-item", ".wd-product", ".product-item"
    ]
    for page in range(1, paginas + 1):
        page_url = url if page == 1 else f"{url.rstrip('/')}/page/{page}/"
        try:
            r    = requests.get(page_url, headers=HEADERS, timeout=30)
            soup = BeautifulSoup(r.text, "html.parser")
            cards = []
            for sel in selectors:
                cards.extend(soup.select(sel))
            if not cards:
                break
            for c in cards:
                try:
                    nombre_el = c.select_one(
                        "h2, h3, h4, .product-title, "
                        ".woocommerce-loop-product__title, .wd-entities-title"
                    )
                    precio_el = c.select_one(".price, .amount, bdi")
                    if not nombre_el or not precio_el:
                        continue
                    nombre     = nombre_el.get_text(strip=True)
                    precio_txt = precio_el.get_text(" ", strip=True)
                    valor, moneda = limpiar_precio(precio_txt, nombre_tienda)
                    if valor is None or valor < 1:
                        continue
                    link_el = c.select_one("a")
                    link = link_el["href"] if link_el else url
                    productos.append({
                        "nombre":          nombre,
                        "precio_original": precio_txt,
                        "precio":          valor,
                        "moneda":          moneda,
                        "precio_uyu":      convertir_uyu(valor, moneda),
                        "link":            link
                    })
                except:
                    continue
        except Exception as e:
            print(f"   ERROR: {e}")
            break
    return productos

def scrape_tiendanube(url, nombre_tienda, paginas=PAGINAS):
    productos = []
    for page in range(1, paginas + 1):
        page_url = url if page == 1 else f"{url}?mpage={page}"
        try:
            r    = requests.get(page_url, headers=HEADERS, timeout=30)
            soup = BeautifulSoup(r.text, "html.parser")
            cards = soup.select(".js-item-product")
            if not cards:
                break
            for c in cards:
                try:
                    nombre     = c.select_one(".js-item-name").get_text(strip=True)
                    precio_txt = c.select_one(".js-price-display").get_text(strip=True)
                    valor, moneda = limpiar_precio(precio_txt, nombre_tienda)
                    if valor is None:
                        continue
                    link = c.select_one("a")["href"]
                    productos.append({
                        "nombre":          nombre,
                        "precio_original": precio_txt,
                        "precio":          valor,
                        "moneda":          moneda,
                        "precio_uyu":      convertir_uyu(valor, moneda),
                        "link":            link
                    })
                except:
                    continue
        except:
            break
    return productos

def scrape_tiendanegocio(url, nombre_tienda, paginas=PAGINAS):
    productos = []
    if not url.startswith("http"):
        url = "https://" + url
    for page in range(1, paginas + 1):
        page_url = url if page == 1 else f"{url}?pagina={page}"
        try:
            r    = requests.get(page_url, headers=HEADERS, timeout=30)
            soup = BeautifulSoup(r.text, "html.parser")
            for c in soup.select(".product-item, .item-producto, article.producto, .card"):
                try:
                    nombre_el  = c.select_one("h2, h3, h4, .product-name, .nombre")
                    precio_el  = c.select_one(".price, .precio, .product-price")
                    if not nombre_el or not precio_el:
                        continue
                    nombre     = nombre_el.get_text(strip=True)
                    precio_txt = precio_el.get_text(strip=True)
                    valor, moneda = limpiar_precio(precio_txt, nombre_tienda)
                    if valor is None:
                        continue
                    link_el = c.select_one("a")
                    link = link_el["href"] if link_el else url
                    productos.append({
                        "nombre":          nombre,
                        "precio_original": precio_txt,
                        "precio":          valor,
                        "moneda":          moneda,
                        "precio_uyu":      convertir_uyu(valor, moneda),
                        "link":            link
                    })
                except:
                    continue
        except Exception as e:
            print(f"   ERROR TiendaNegocio: {e}")
            break
    return productos

# =========================================================
# SCRAPER URUGUAYDIGITAL
# =========================================================
async def scrape_uruguaydigital(url, nombre_tienda, paginas=PAGINAS):
    productos = []
    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage"]
        )
        page = await browser.new_page(
            user_agent=HEADERS["User-Agent"],
            locale="es-419"
        )
        for num in range(1, paginas + 1):
            page_url = url if num == 1 else f"{url}?pagina={num}"
            try:
                await page.goto(page_url, wait_until="domcontentloaded", timeout=40000)
                await page.wait_for_timeout(3000)
                html = await page.content()
                soup = BeautifulSoup(html, "html.parser")
                cards = soup.select("li.d-flex.no-wrap.justify-content-start.align-items-center")
                if not cards:
                    break
                for c in cards:
                    try:
                        texto  = c.get_text("\n", strip=True)
                        lineas = [l.strip() for l in texto.split("\n") if l.strip()]
                        if len(lineas) < 2:
                            continue
                        nombre    = ""
                        precio_txt = ""
                        for linea in lineas:
                            if re.search(r"\$[\d.,]+", linea):
                                precio_txt = linea
                                break
                            else:
                                nombre = linea
                        if not nombre or not precio_txt:
                            continue
                        valor, moneda = limpiar_precio(precio_txt, nombre_tienda)
                        if valor is None or valor < 1:
                            continue
                        link_el = c.select_one("a")
                        link = link_el["href"] if link_el else url
                        if link and not link.startswith("http"):
                            link = "https://juegosdigitalesuruguay.com" + link
                        productos.append({
                            "nombre":          nombre,
                            "precio_original": precio_txt,
                            "precio":          valor,
                            "moneda":          moneda,
                            "precio_uyu":      convertir_uyu(valor, moneda),
                            "link":            link
                        })
                    except:
                        continue
            except Exception as e:
                print(f"   ERROR UruguayDigital p{num}: {e}")
                break
        await browser.close()
    return productos

# =========================================================
# SCRAPER JDP4P5
# =========================================================
async def scrape_jdp4p5(url, nombre_tienda, paginas=PAGINAS):
    productos = []
    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage"]
        )
        page = await browser.new_page(
            user_agent=HEADERS["User-Agent"],
            locale="es-419"
        )
        for num in range(1, paginas + 1):
            page_url = url if num == 1 else f"{url}?page={num}"
            try:
                await page.goto(page_url, wait_until="domcontentloaded", timeout=40000)
                await page.wait_for_timeout(3000)
                html  = await page.content()
                soup  = BeautifulSoup(html, "html.parser")
                cards = soup.select("div.js-item-product")
                if not cards:
                    break
                for c in cards:
                    try:
                        link_el   = c.select_one("a.item-link")
                        nombre_el = c.select_one(".js-item-name, h2, h3")
                        if link_el and not nombre_el:
                            nombre = link_el.get_text("\n", strip=True).split("\n")[0].strip()
                        elif nombre_el:
                            nombre = nombre_el.get_text(strip=True)
                        else:
                            continue
                        precio_el = c.select_one("span.js-price-display")
                        if not precio_el:
                            continue
                        precio_txt    = precio_el.get_text(strip=True)
                        precio_limpio = re.sub(r"[^\d,]", "", precio_txt).replace(",", ".")
                        try:
                            valor = float(precio_limpio)
                        except:
                            continue
                        moneda = MONEDA_TIENDA.get(nombre_tienda, "ARS")
                        if valor < 1:
                            continue
                        link = link_el["href"] if link_el else url
                        if link and not link.startswith("http"):
                            link = "https://juegosdigitalesps4ps5.com" + link
                        productos.append({
                            "nombre":          nombre,
                            "precio_original": precio_txt,
                            "precio":          valor,
                            "moneda":          moneda,
                            "precio_uyu":      convertir_uyu(valor, moneda),
                            "link":            link
                        })
                    except:
                        continue
            except Exception as e:
                print(f"   ERROR JDP4P5 p{num}: {e}")
                break
        await browser.close()
    return productos

# =========================================================
# SCRAPER ZONADIGITAL
# =========================================================
async def scrape_zonadigital(url, nombre_tienda, paginas=PAGINAS):
    productos = []
    if not url.startswith("http"):
        url = "https://" + url
    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage"]
        )
        page = await browser.new_page(
            user_agent=HEADERS["User-Agent"],
            locale="es-419"
        )
        for num in range(1, paginas + 1):
            page_url = url if num == 1 else f"{url}?pagina={num}"
            try:
                await page.goto(page_url, wait_until="networkidle", timeout=50000)
                await page.wait_for_timeout(4000)
                for _ in range(4):
                    await page.mouse.wheel(0, 4000)
                    await page.wait_for_timeout(1500)
                html  = await page.content()
                soup  = BeautifulSoup(html, "html.parser")
                cards = soup.select("div.item-gift__content")
                if not cards:
                    break
                for c in cards:
                    try:
                        nombre = ""
                        for child in c.children:
                            txt = child.get_text(strip=True) if hasattr(child, "get_text") else str(child).strip()
                            if txt and not txt.startswith("De") and "$" not in txt:
                                nombre = txt
                                break
                        if not nombre:
                            continue
                        precio_el = c.select_one("span:not([class])")
                        if not precio_el:
                            for sp in c.select("span"):
                                if "$" in sp.get_text():
                                    precio_el = sp
                                    break
                        if not precio_el:
                            continue
                        precio_txt    = precio_el.get_text(strip=True)
                        precio_limpio = re.sub(r"[^\d]", "", precio_txt)
                        try:
                            valor = float(precio_limpio)
                        except:
                            continue
                        moneda = MONEDA_TIENDA.get(nombre_tienda, "UYU")
                        if valor < 1:
                            continue
                        link_el = c.select_one("a")
                        link = link_el["href"] if link_el else url
                        if link and not link.startswith("http"):
                            link = "https://zonadigitalmd.com" + link
                        productos.append({
                            "nombre":          nombre,
                            "precio_original": precio_txt,
                            "precio":          valor,
                            "moneda":          moneda,
                            "precio_uyu":      convertir_uyu(valor, moneda),
                            "link":            link
                        })
                    except:
                        continue
            except Exception as e:
                print(f"   ERROR ZonaDigital p{num}: {e}")
                break
        await browser.close()
    return productos

# =========================================================
# CATALOGO COMPLETO DE TIENDAS
# =========================================================
CATALOGOS = {
    # ---- ALPHA (Shopify) ----
    "Alpha PS5 Primaria":       {"tipo": "shopify",       "url": "https://alphajuegosdigitales.com/collections/ps5-principal"},
    "Alpha PS5 Secundaria":     {"tipo": "shopify",       "url": "https://alphajuegosdigitales.com/collections/juegos-ps5"},
    "Alpha PS4 Primaria":       {"tipo": "shopify",       "url": "https://alphajuegosdigitales.com/collections/ps4-principal"},
    "Alpha PS4 Secundaria":     {"tipo": "shopify",       "url": "https://alphajuegosdigitales.com/collections/juegos-ps4"},
    "Alpha Xbox":               {"tipo": "shopify",       "url": "https://alphajuegosdigitales.com/collections/juegos-xbox"},
    # ---- DIGITALWORLD (WooCommerce) ----
    "DigitalWorld PS4":         {"tipo": "woocommerce",   "url": "https://digitalworldpsn.com/es/juegos-digitales-ps4/?v=1b23f8a4c97c"},
    "DigitalWorld PS5":         {"tipo": "woocommerce",   "url": "https://digitalworldpsn.com/es/juegos-digitales-ps5/?v=1b23f8a4c97c"},
    "DigitalWorld VR":          {"tipo": "woocommerce",   "url": "https://digitalworldpsn.com/es/juegos-ps-vr-vr2/?v=1b23f8a4c97c"},
    "DigitalWorld Xbox":        {"tipo": "woocommerce",   "url": "https://digitalworldpsn.com/es/juegos-digitales-xbox/?v=1b23f8a4c97c"},
    "DigitalWorld Switch":      {"tipo": "woocommerce",   "url": "https://digitalworldpsn.com/es/juegos-digitales-switch/?v=1b23f8a4c97c"},
    # ---- DIXGAMER (WooCommerce) ----
    "Dix Fortnite":             {"tipo": "woocommerce",   "url": "https://dixgamer.com/categoria-producto/tarjetas/fortnite/"},
    "Dix PS5":                  {"tipo": "woocommerce",   "url": "https://dixgamer.com/categoria-producto/juegos/ps5/"},
    "Dix PS4":                  {"tipo": "woocommerce",   "url": "https://dixgamer.com/categoria-producto/juegos/ps4/"},
    "Dix PS3":                  {"tipo": "woocommerce",   "url": "https://dixgamer.com/categoria-producto/juegos/ps3/"},
    "Dix FC Points":            {"tipo": "woocommerce",   "url": "https://dixgamer.com/categoria-producto/tarjetas/fc-points/"},
    "Dix Nintendo":             {"tipo": "woocommerce",   "url": "https://dixgamer.com/categoria-producto/tarjetas/nintendo/"},
    "Dix Playstation":          {"tipo": "woocommerce",   "url": "https://dixgamer.com/categoria-producto/tarjetas/ps/"},
    "Dix Razer":                {"tipo": "woocommerce",   "url": "https://dixgamer.com/categoria-producto/tarjetas/razer-gold/"},
    "Dix Steam":                {"tipo": "woocommerce",   "url": "https://dixgamer.com/categoria-producto/tarjetas/steam/"},
    "Dix Xbox":                 {"tipo": "woocommerce",   "url": "https://dixgamer.com/categoria-producto/tarjetas/xbox/"},
    # ---- ESTACIONPLAY (Tiendanube) — argentina → ARS ----
    "EstacionPlay PS4":         {"tipo": "tiendanube",    "url": "https://estacionplay.com/videojuegos/playstation-4/"},
    "EstacionPlay PS5":         {"tipo": "tiendanube",    "url": "https://estacionplay.com/videojuegos/playstation-5/"},
    "EstacionPlay Memb":        {"tipo": "tiendanube",    "url": "https://estacionplay.com/videojuegos/psn-card-y-plus/"},
    # ---- GAMERLAB (Shopify) ----
    "GamerLab PS4":             {"tipo": "shopify",       "url": "https://juegosdigitalesgamerlab.com/collections/frontpage"},
    "GamerLab PS5":             {"tipo": "shopify",       "url": "https://juegosdigitalesgamerlab.com/collections/juegos-ps5"},
    "GamerLab EA Play":         {"tipo": "shopify",       "url": "https://juegosdigitalesgamerlab.com/collections/membresia-ea-play"},
    "GamerLab PSN":             {"tipo": "shopify",       "url": "https://juegosdigitalesgamerlab.com/collections/play-station-plus"},
    "GamerLab All":             {"tipo": "shopify",       "url": "https://juegosdigitalesgamerlab.com/collections/all"},
    # ---- JUEGOSDIGITALESPS4PS5 (Tiendanube con Playwright) — argentina → ARS ----
    "JDP4P5 PS4":               {"tipo": "jdp4p5",        "url": "https://juegosdigitalesps4ps5.com/juegos-ps4/"},
    "JDP4P5 PS5":               {"tipo": "jdp4p5",        "url": "https://juegosdigitalesps4ps5.com/juegos-ps5/"},
    "JDP4P5 Switch":            {"tipo": "jdp4p5",        "url": "https://juegosdigitalesps4ps5.com/juegos-nintendo/"},
    "JDP4P5 PS Plus":           {"tipo": "jdp4p5",        "url": "https://juegosdigitalesps4ps5.com/membresias-ps-plus/"},
    "JDP4P5 GamePass":          {"tipo": "jdp4p5",        "url": "https://juegosdigitalesps4ps5.com/membresias-game-pass/"},
    "JDP4P5 Nintendo Online":   {"tipo": "jdp4p5",        "url": "https://juegosdigitalesps4ps5.com/suscripcion-nintendo-online/"},
    # ---- JUEGOSDIGITALESURUGUAY (sistema custom con Playwright) ----
    "UruguayDigital PS3":       {"tipo": "uruguaydigital", "url": "https://juegosdigitalesuruguay.com/categorias/juegos-digitales-ps3"},
    "UruguayDigital PS4":       {"tipo": "uruguaydigital", "url": "https://juegosdigitalesuruguay.com/categorias/juegos-digitales-ps4"},
    "UruguayDigital PS5":       {"tipo": "uruguaydigital", "url": "https://juegosdigitalesuruguay.com/categorias/juegos-digitales-ps5"},
    "UruguayDigital Switch1":   {"tipo": "uruguaydigital", "url": "https://juegosdigitalesuruguay.com/categorias/nintendo-switch/juegos-nintendo-switch"},
    "UruguayDigital Switch2":   {"tipo": "uruguaydigital", "url": "https://www.juegosdigitalesuruguay.com/categorias/nintendo-switch-2"},
    "UruguayDigital Xbox One":  {"tipo": "uruguaydigital", "url": "https://juegosdigitalesuruguay.com/categorias/xbox/juegos-xbox-one"},
    "UruguayDigital Xbox XS":   {"tipo": "uruguaydigital", "url": "https://juegosdigitalesuruguay.com/categorias/xbox/juegos-xbox-series-xs"},
    "UruguayDigital PSN Plus":  {"tipo": "uruguaydigital", "url": "https://juegosdigitalesuruguay.com/categorias/cuenta-psn-plus-global"},
    "UruguayDigital Nintendo":  {"tipo": "uruguaydigital", "url": "https://juegosdigitalesuruguay.com/categorias/nintendo-membresia"},
    "UruguayDigital GamePass":  {"tipo": "uruguaydigital", "url": "https://juegosdigitalesuruguay.com/categorias/membresias-xbox"},
    # ---- MDQ STORE (WooCommerce) — argentina → ARS ----
    "MDQ PS5":                  {"tipo": "woocommerce",   "url": "https://mdqstore.com/categoria-producto/juegos-ps5/"},
    "MDQ PS4":                  {"tipo": "woocommerce",   "url": "https://mdqstore.com/categoria-producto/juegos-ps4/"},
    "MDQ PS3":                  {"tipo": "woocommerce",   "url": "https://mdqstore.com/categoria-producto/juegos-ps3/"},
    "MDQ PS Plus":              {"tipo": "woocommerce",   "url": "https://mdqstore.com/categoria-producto/playstation-plus/"},
    "MDQ PSN Card":             {"tipo": "woocommerce",   "url": "https://mdqstore.com/categoria-producto/psn-card/"},
    "MDQ Fortnite":             {"tipo": "woocommerce",   "url": "https://mdqstore.com/categoria-producto/fornite/"},
    # ---- PLAYX DIGITAL (WooCommerce) — argentina → ARS ----
    "PlayX PS4":                {"tipo": "woocommerce",   "url": "https://playxdigital.com/categoria-producto/psn/ps4"},
    "PlayX PS5":                {"tipo": "woocommerce",   "url": "https://playxdigital.com/categoria-producto/psn/ps5"},
    "PlayX Xbox One":           {"tipo": "woocommerce",   "url": "https://playxdigital.com/categoria-producto/xbox/xbox-one"},
    "PlayX Xbox XS":            {"tipo": "woocommerce",   "url": "https://playxdigital.com/categoria-producto/xbox/xbox-series-x-s"},
    "PlayX Switch":             {"tipo": "woocommerce",   "url": "https://playxdigital.com/categoria-producto/nintendo/nintendo-switch"},
    "PlayX Switch 2":           {"tipo": "woocommerce",   "url": "https://playxdigital.com/categoria-producto/nintendo/switch2/"},
    "PlayX Steam":              {"tipo": "woocommerce",   "url": "https://playxdigital.com/categoria-producto/steam"},
    "PlayX PSN Plus":           {"tipo": "woocommerce",   "url": "https://playxdigital.com/membresias/psn-plus"},
    "PlayX GamePass":           {"tipo": "woocommerce",   "url": "https://playxdigital.com/membresias/game-pass-ultimate"},
    "PlayX Switch Online":      {"tipo": "woocommerce",   "url": "https://playxdigital.com/membresias/swicth-online/"},
    # ---- PORTAL GAMES (WooCommerce / Playwright) — argentina → ARS ----
    "PortalGames PS3":          {"tipo": "playwright",    "url": "https://portalgames.com.ar/product-category/juegos-ps3/"},
    "PortalGames PS4":          {"tipo": "playwright",    "url": "https://portalgames.com.ar/product-category/juegos-ps4/"},
    "PortalGames PS5":          {"tipo": "playwright",    "url": "https://portalgames.com.ar/product-category/juegos-ps5/"},
    "PortalGames Switch":       {"tipo": "playwright",    "url": "https://portalgames.com.ar/product-category/nintendo-switch/"},
    # ---- TODO DIGITAL SHOP (WooCommerce / Playwright) — argentina → ARS ----
    "TodoDigital PS4":          {"tipo": "playwright",    "url": "https://tododigitalshop.com/juegos-digitales-ps4/"},
    "TodoDigital PS5":          {"tipo": "playwright",    "url": "https://tododigitalshop.com/juegos-digitales-ps5/"},
    "TodoDigital Switch":       {"tipo": "playwright",    "url": "https://tododigitalshop.com/juegos-digitales-nintendo-switch/"},
    # ---- URUGUAY JUEGOS DIGITALES (WooCommerce) ----
    "UYJuegos PS4":             {"tipo": "woocommerce",   "url": "https://uruguayjuegosdigitales.com/product-category/juegos-digitales-ps4/"},
    "UYJuegos PS4 VR":          {"tipo": "woocommerce",   "url": "https://uruguayjuegosdigitales.com/product-category/juegos-digitales-ps4/vr/"},
    "UYJuegos PS5":             {"tipo": "woocommerce",   "url": "https://uruguayjuegosdigitales.com/product-category/juegos-digitales-ps5/"},
    "UYJuegos PC":              {"tipo": "woocommerce",   "url": "https://uruguayjuegosdigitales.com/product-category/juegos-digitales-pc/todos-los-juegos-pc/"},
    "UYJuegos Xbox XS":         {"tipo": "woocommerce",   "url": "https://uruguayjuegosdigitales.com/product-category/juegos-digitales-xbox/juegos-digitales-xbox-series-x-s/"},
    "UYJuegos Xbox One":        {"tipo": "woocommerce",   "url": "https://uruguayjuegosdigitales.com/product-category/juegos-digitales-xbox/juegos-digitales-xbox-one/"},
    "UYJuegos Nintendo":        {"tipo": "woocommerce",   "url": "https://uruguayjuegosdigitales.com/product-category/juegos-nintendo/juegos-nintendo-switch/"},
    "UYJuegos Nintendo2":       {"tipo": "woocommerce",   "url": "https://uruguayjuegosdigitales.com/product-category/nintendo-switch-2"},
    "UYJuegos GiftCard Amazon": {"tipo": "woocommerce",   "url": "https://uruguayjuegosdigitales.com/product-category/gift-cards/tarjetas-de-regalo-amazon/"},
    "UYJuegos GiftCard iTunes": {"tipo": "woocommerce",   "url": "https://uruguayjuegosdigitales.com/product-category/gift-cards/itunes/"},
    # ---- WEB-GAME.NET (WooCommerce) ----
    "WebGame PS4":              {"tipo": "woocommerce",   "url": "https://web-game.net/categoria/juegos-ps4/"},
    "WebGame PS5":              {"tipo": "woocommerce",   "url": "https://web-game.net/categoria/juegos-ps5/"},
    "WebGame GiftCards":        {"tipo": "woocommerce",   "url": "https://web-game.net/categoria/gift-card/"},
    # ---- ZONADIGITALMD (Angular — requiere Playwright) ----
    "ZonaDigital PS4":          {"tipo": "zonadigital",   "url": "zonadigitalmd.com/productos/juegosps4"},
    "ZonaDigital PS5":          {"tipo": "zonadigital",   "url": "zonadigitalmd.com/productos/juegosps5"},
    "ZonaDigital PC Codigo":    {"tipo": "zonadigital",   "url": "zonadigitalmd.com/productos/juegospccodigo"},
    "ZonaDigital PC Cuenta":    {"tipo": "zonadigital",   "url": "zonadigitalmd.com/productos/juegospccuenta"},
    "ZonaDigital Xbox":         {"tipo": "zonadigital",   "url": "zonadigitalmd.com/productos/juegosxbox"},
    "ZonaDigital Apex":         {"tipo": "zonadigital",   "url": "zonadigitalmd.com/productos/monedasapexlegends"},
    "ZonaDigital Fortnite VB":  {"tipo": "zonadigital",   "url": "zonadigitalmd.com/productos/fortnitevbucks"},
    "ZonaDigital Fortnite":     {"tipo": "zonadigital",   "url": "zonadigitalmd.com/productos/cargafortnite"},
    "ZonaDigital GiftCard PSN": {"tipo": "zonadigital",   "url": "zonadigitalmd.com/productos/giftcardps3ps4ps5"},
    "ZonaDigital GiftCard Steam": {"tipo": "zonadigital", "url": "zonadigitalmd.com/productos/giftcardsteam"},
    "ZonaDigital Roblox":       {"tipo": "zonadigital",   "url": "zonadigitalmd.com/productos/giftcardroblox"},
    "ZonaDigital GiftCard Xbox":{"tipo": "zonadigital",   "url": "zonadigitalmd.com/productos/giftcardxbox"},
    "ZonaDigital BattleNet":    {"tipo": "zonadigital",   "url": "zonadigitalmd.com/productos/giftcardbattlenet"},
    "ZonaDigital Riot":         {"tipo": "zonadigital",   "url": "zonadigitalmd.com/productos/giftcardriotlatam"},
    "ZonaDigital FreeFire":     {"tipo": "zonadigital",   "url": "zonadigitalmd.com/productos/giftcardfreefire"},
    "ZonaDigital Google Play":  {"tipo": "zonadigital",   "url": "zonadigitalmd.com/productos/giftcardgoogleplayusa"},
    "ZonaDigital Apple":        {"tipo": "zonadigital",   "url": "zonadigitalmd.com/productos/giftcardappleusa"},
    "ZonaDigital Switch":       {"tipo": "zonadigital",   "url": "zonadigitalmd.com/productos/nintendoswitch"},
    "ZonaDigital Switch USA":   {"tipo": "zonadigital",   "url": "zonadigitalmd.com/productos/giftcardusa"},
    "ZonaDigital GamePass":     {"tipo": "zonadigital",   "url": "zonadigitalmd.com/productos/suscripcioncodigo"},
    "ZonaDigital EA Play":      {"tipo": "zonadigital",   "url": "zonadigitalmd.com/productos/eaplaycodigo"},
}

# =========================================================
# PAGINAS ENEBA
# =========================================================
PAGINAS_ENEBA = [
    ("FreeFire",      "https://www.eneba.com/latam/top-up-free-fire-diamonds-global",                     "producto"),
    ("Genshin",       "https://www.eneba.com/latam/top-up-genshin-impact-genesis-crystals-latin-america", "producto"),
    ("Robux",         "https://www.eneba.com/latam/store/all?text=robux",                                 "tienda"),
    ("PUBG",          "https://www.eneba.com/latam/store/all?text=pubg+uc",                               "tienda"),
    ("FC Points",     "https://www.eneba.com/latam/store/fc-points",                                      "tienda"),
    ("FUT Points",    "https://www.eneba.com/latam/store/game-points-fut",                                "tienda"),
    ("GTA Shark",     "https://www.eneba.com/latam/store/gta-shark-cards",                                "tienda"),
    ("Valorant",      "https://www.eneba.com/latam/store/riot-valorant-points",                           "tienda"),
    ("COD Points",    "https://www.eneba.com/latam/store/cod-points",                                     "tienda"),
    ("Fortnite",      "https://www.eneba.com/latam/store/fortnite-v-bucks-gift-cards",                    "tienda"),
    ("Xbox",          "https://www.eneba.com/latam/store/xbox-gift-cards",                                "tienda"),
    ("Xbox Points",   "https://www.eneba.com/latam/store/xbox-game-points",                               "tienda"),
    ("Xbox GamePass", "https://www.eneba.com/latam/store/xbox-game-pass",                                 "tienda"),
    ("Xbox Games",    "https://www.eneba.com/latam/store/xbox-games",                                     "tienda"),
    ("PlayStation",   "https://www.eneba.com/latam/store/psn-games",                                      "tienda"),
    ("PSN GiftCards", "https://www.eneba.com/latam/store/psn-gift-cards",                                 "tienda"),
    ("PSN Plus",      "https://www.eneba.com/latam/store/psn-subscriptions",                              "tienda"),
    ("Nintendo",      "https://www.eneba.com/latam/store/nintendo-games",                                 "tienda"),
    ("Nintendo Gift", "https://www.eneba.com/latam/store/nintendo-gift-cards",                            "tienda"),
    ("Nintendo Subs", "https://www.eneba.com/latam/store/nintendo-subscriptions",                         "tienda"),
    ("Steam",         "https://www.eneba.com/latam/store/steam-games",                                    "tienda"),
    ("Steam Wallet",  "https://www.eneba.com/latam/store/steam-gift-cards",                               "tienda"),
    ("EA Play",       "https://www.eneba.com/latam/store/ea-play",                                        "tienda"),
    ("Origin",        "https://www.eneba.com/latam/store/origin-games",                                   "tienda"),
    ("Ubisoft",       "https://www.eneba.com/latam/store/uplay-games",                                    "tienda"),
    ("Epic Games",    "https://www.eneba.com/latam/store/epic-games",                                     "tienda"),
    ("GOG",           "https://www.eneba.com/latam/store/gog-games",                                      "tienda"),
    ("BattleNet",     "https://www.eneba.com/latam/store/battle-net-games",                               "tienda"),
    ("BattleNetPts",  "https://www.eneba.com/latam/store/battle-net-game-points",                         "tienda"),
    ("Google Play",   "https://www.eneba.com/latam/store/google-play-gift-cards",                         "tienda"),
    ("Razer Gold",    "https://www.eneba.com/latam/store/razer-gold-gift-cards",                          "tienda"),
    ("Discord",       "https://www.eneba.com/latam/store/discord-gift-cards",                             "tienda"),
    ("Twitch",        "https://www.eneba.com/latam/store/twitch-gift-cards",                              "tienda"),
    ("Blizzard",      "https://www.eneba.com/latam/store/blizzard-gift-card",                             "tienda"),
]

# =========================================================
# MAIN ASYNC
# =========================================================
async def main():
    todos = []

    # --------------------------------------------------
    # Playwright genérico para WooCommerce con JS
    # --------------------------------------------------
    async def scrape_playwright(url, nombre_tienda, paginas=PAGINAS):
        prods = []
        selectors = [
            "li.product", "div.product", ".product-small",
            ".type-product", ".product-grid-item", ".wd-product", ".product-item"
        ]
        async with async_playwright() as pw:
            browser = await pw.chromium.launch(
                headless=True,
                args=["--no-sandbox", "--disable-dev-shm-usage"]
            )
            page = await browser.new_page(
                user_agent=HEADERS["User-Agent"],
                locale="es-419"
            )
            for num in range(1, paginas + 1):
                page_url = url if num == 1 else f"{url.rstrip('/')}/page/{num}/"
                try:
                    await page.goto(page_url, wait_until="domcontentloaded", timeout=40000)
                    await page.wait_for_timeout(3000)
                    html  = await page.content()
                    soup  = BeautifulSoup(html, "html.parser")
                    cards = []
                    for sel in selectors:
                        cards.extend(soup.select(sel))
                    for c in cards:
                        try:
                            nombre_el  = c.select_one(
                                "h2, h3, h4, .product-title, "
                                ".woocommerce-loop-product__title, .wd-entities-title"
                            )
                            precio_el  = c.select_one(".price, .amount, bdi")
                            if not nombre_el or not precio_el:
                                continue
                            nombre     = nombre_el.get_text(strip=True)
                            precio_txt = precio_el.get_text(" ", strip=True)
                            valor, moneda = limpiar_precio(precio_txt, nombre_tienda)
                            if valor is None or valor < 1:
                                continue
                            link_el = c.select_one("a")
                            link = link_el["href"] if link_el else url
                            prods.append({
                                "nombre":          nombre,
                                "precio_original": precio_txt,
                                "precio":          valor,
                                "moneda":          moneda,
                                "precio_uyu":      convertir_uyu(valor, moneda),
                                "link":            link
                            })
                        except:
                            continue
                except Exception as e:
                    print(f"   ERROR Playwright p{num}: {e}")
            await browser.close()
        return prods

    # --------------------------------------------------
    # PARTE 1: CATALOGO MAESTRO
    # --------------------------------------------------
    print("\n" + "=" * 60)
    print("SCRAPEANDO CATALOGO DE TIENDAS")
    print("=" * 60)

    for nombre_tienda, cfg in CATALOGOS.items():
        print(f"\n  {nombre_tienda}...", end=" ", flush=True)
        try:
            tipo = cfg["tipo"]
            url  = cfg["url"]
            if tipo == "shopify":
                productos = scrape_shopify(url, nombre_tienda)
            elif tipo == "tiendanube":
                productos = scrape_tiendanube(url, nombre_tienda)
            elif tipo == "tiendanegocio":
                productos = scrape_tiendanegocio(url, nombre_tienda)
            elif tipo == "playwright":
                productos = await scrape_playwright(url, nombre_tienda)
            elif tipo == "uruguaydigital":
                productos = await scrape_uruguaydigital(url, nombre_tienda)
            elif tipo == "jdp4p5":
                productos = await scrape_jdp4p5(url, nombre_tienda)
            elif tipo == "zonadigital":
                productos = await scrape_zonadigital(url, nombre_tienda)
            else:
                productos = scrape_woocommerce(url, nombre_tienda)
                if len(productos) == 0:
                    print("(retry playwright)", end=" ", flush=True)
                    productos = await scrape_playwright(url, nombre_tienda)

            print(f"{len(productos)} productos")
            for p in productos:
                p["tienda"]             = nombre_tienda
                p["fuente"]             = tipo.capitalize()
                p["nombre_normalizado"] = normalizar_nombre(p["nombre"])
                p["plataforma"]         = detectar_plataforma(p["nombre"], nombre_tienda)
                todos.append(p)
        except Exception as e:
            print(f"ERROR: {e}")

    # --------------------------------------------------
    # PARTE 2: ENEBA
    # --------------------------------------------------
    print("\n" + "=" * 60)
    print("SCRAPEANDO ENEBA")
    print("=" * 60)

    BASE_ENEBA   = "https://www.eneba.com"
    vistos_eneba = set()

    async def cerrar_popups(page):
        for texto in ["Aceptar todo", "Accept all", "Entendido"]:
            try:
                b = await page.query_selector(f"button:has-text('{texto}')")
                if b:
                    await b.click()
                    await page.wait_for_timeout(800)
            except:
                pass
        try:
            await page.keyboard.press("Escape")
        except:
            pass

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage"]
        )
        for categoria, url, tipo in PAGINAS_ENEBA:
            nombre_tienda_eneba = f"Eneba - {categoria}"
            print(f"  {nombre_tienda_eneba}...", end=" ", flush=True)
            try:
                page = await browser.new_page(
                    viewport={"width": 1366, "height": 768},
                    user_agent=HEADERS["User-Agent"],
                    locale="es-419",
                    extra_http_headers={"Accept-Language": "es-419,es;q=0.9"}
                )
                await page.goto(url, wait_until="domcontentloaded", timeout=30000)
                await page.wait_for_timeout(3000)
                await cerrar_popups(page)
                await page.wait_for_timeout(1500)
                for _ in range(4):
                    await page.mouse.wheel(0, 4000)
                    await page.wait_for_timeout(1200)
                html = await page.content()
                soup = BeautifulSoup(html, "html.parser")
                count = 0

                if tipo == "producto":
                    cards = soup.find_all("div", class_="vNgEk7")
                    for card in cards:
                        try:
                            lineas = [l.strip() for l in card.get_text("\n").split("\n") if l.strip()]
                            nombre = lineas[0] if lineas else ""
                            precio = next(
                                (extraer_precio_eneba(l) for l in lineas if extraer_precio_eneba(l)), ""
                            )
                            if not nombre or not precio:
                                continue
                            clave = (nombre, precio)
                            if clave in vistos_eneba:
                                continue
                            vistos_eneba.add(clave)
                            valor, moneda = limpiar_precio(precio, "Eneba")
                            todos.append({
                                "nombre":             nombre,
                                "precio_original":    precio,
                                "precio":             valor,
                                "moneda":             moneda,
                                "precio_uyu":         convertir_uyu(valor, moneda),
                                "link":               url,
                                "tienda":             nombre_tienda_eneba,
                                "fuente":             "Eneba",
                                "nombre_normalizado": limpiar_nombre_eneba(nombre),
                                "plataforma":         detectar_plataforma(nombre, categoria)
                            })
                            count += 1
                        except:
                            pass

                elif tipo == "tienda":
                    cards = soup.find_all("div", class_="b3POZC")
                    for card in cards:
                        try:
                            a      = card.find("a", href=True)
                            if not a and card.parent:
                                a = card.parent.find("a", href=True)
                            nombre = ""
                            link   = url
                            if a:
                                nombre = a.get("title", "").strip()
                                if not nombre:
                                    img = a.find("img")
                                    if img:
                                        nombre = img.get("alt", "").strip()
                                href = a.get("href", "")
                                link = BASE_ENEBA + href if href.startswith("/") else href
                            nombre = limpiar_nombre_eneba(nombre)
                            precio = next(
                                (
                                    extraer_precio_eneba(l.strip())
                                    for l in card.get_text("\n").split("\n")
                                    if extraer_precio_eneba(l.strip())
                                ),
                                ""
                            )
                            if not nombre or not precio:
                                continue
                            clave = (nombre, precio)
                            if clave in vistos_eneba:
                                continue
                            vistos_eneba.add(clave)
                            valor, moneda = limpiar_precio(precio, "Eneba")
                            todos.append({
                                "nombre":             nombre,
                                "precio_original":    precio,
                                "precio":             valor,
                                "moneda":             moneda,
                                "precio_uyu":         convertir_uyu(valor, moneda),
                                "link":               link,
                                "tienda":             nombre_tienda_eneba,
                                "fuente":             "Eneba",
                                "nombre_normalizado": limpiar_nombre_eneba(nombre),
                                "plataforma":         detectar_plataforma(nombre, categoria)
                            })
                            count += 1
                        except:
                            pass

                print(f"{count} productos")
                await page.close()
            except Exception as e:
                print(f"ERROR: {e}")
        await browser.close()

    # --------------------------------------------------
    # PARTE 3: ARMAR Y EXPORTAR EXCEL
    # --------------------------------------------------
    print("\n" + "=" * 60)
    print("ARMANDO EXCEL")
    print("=" * 60)

    df = pd.DataFrame(todos)
    print(f"Total bruto: {len(df)} productos")
    if len(df) == 0:
        print("Sin datos. Saliendo.")
        return

    df = df[
        (df["precio_uyu"].notna()) &
        (df["precio_uyu"] > 5) &
        (df["precio_uyu"] < 500000)
    ].copy()
    df = df.sort_values("precio_uyu")

    idx_mejores = df.groupby("nombre_normalizado")["precio_uyu"].idxmin()
    mejores     = df.loc[idx_mejores].sort_values("precio_uyu").copy()

    COLS_DISPLAY = [
        "fuente", "tienda", "nombre", "nombre_normalizado",
        "precio_original", "moneda", "precio_uyu", "plataforma", "link"
    ]
    COLS_LABELS = [
        "Fuente", "Tienda", "Nombre", "Nombre Limpio",
        "Precio Original", "Moneda", "Precio UYU", "Plataforma", "Link"
    ]
    COL_MAP = dict(zip(COLS_DISPLAY, COLS_LABELS))
    ARCHIVO = "catalogo_completo_abc_gaming.xlsx"

    with pd.ExcelWriter(ARCHIVO, engine="openpyxl") as writer:
        df[COLS_DISPLAY].rename(columns=COL_MAP).to_excel(
            writer, sheet_name="TODO", index=False
        )
        mejores[COLS_DISPLAY].rename(columns=COL_MAP).to_excel(
            writer, sheet_name="MEJORES_PRECIOS", index=False
        )
        eneba_df = df[df["fuente"] == "Eneba"]
        if len(eneba_df) > 0:
            eneba_df[COLS_DISPLAY].rename(columns=COL_MAP).to_excel(
                writer, sheet_name="ENEBA", index=False
            )
        for plataforma in sorted(df["plataforma"].unique()):
            temp = df[df["plataforma"] == plataforma]
            if len(temp) == 0:
                continue
            nombre_hoja = plataforma[:31]
            temp[COLS_DISPLAY].rename(columns=COL_MAP).to_excel(
                writer, sheet_name=nombre_hoja, index=False
            )

    # --------------------------------------------------
    # FORMATO VISUAL
    # --------------------------------------------------
    wb          = load_workbook(ARCHIVO)
    HEADER_FILL = PatternFill("solid", start_color="1E1E2E", end_color="1E1E2E")
    HEADER_FONT = Font(bold=True, color="00D4FF", name="Arial", size=10)
    ALT_FILL    = PatternFill("solid", start_color="F5F5F5", end_color="F5F5F5")
    LINK_FONT   = Font(color="0000FF", underline="single")

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if i % 2 == 0:
                for cell in row:
                    cell.fill = ALT_FILL
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        if "Link" in headers:
            col_link = headers.index("Link") + 1
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row, col_link)
                if cell.value and str(cell.value).startswith("http"):
                    cell.hyperlink = cell.value
                    cell.font      = LINK_FONT
        for col_idx, col in enumerate(ws.columns, start=1):
            max_len = max(
                (len(str(c.value)) for c in col if c.value),
                default=10
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 60)
        ws.freeze_panes    = "A2"
        ws.auto_filter.ref = ws.dimensions

    wb.save(ARCHIVO)
    print(f"\nTotal productos finales : {len(df)}")
    print(f"Juegos únicos (mejor precio): {len(mejores)}")
    print(f"Archivo generado: {ARCHIVO}")

# =========================================================
# EJECUTAR
# =========================================================
if __name__ == "__main__":
    asyncio.run(main())
